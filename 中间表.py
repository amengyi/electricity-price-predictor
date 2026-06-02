import os
import numpy as np
import pandas as pd
import psycopg2
from sqlalchemy import create_engine
import warnings

warnings.filterwarnings('ignore')
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)

import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

psycopg2.extensions.register_adapter(np.int64, psycopg2._psycopg.AsIs)

# 源数据库连接 - 从环境变量读取
OUTER_HOST = 'pgm-bp100hm7m50ii06z.pg.rds.aliyuncs.com'
INNER_HOST = 'pgm-bp100hm7m50ii06z.pg.rds.aliyuncs.com'

try:
    NAMESPACE = os.environ["KUBERNETES_NAMESPACE"]
    HOST = INNER_HOST
except KeyError:
    HOST = OUTER_HOST

USER = os.environ.get('SOURCE_DB_USER', 'reader_anhui')
DATABASE = os.environ.get('SOURCE_DB_NAME', 'bopha_anhui')
PASSWORD = os.environ.get('SOURCE_DB_PASSWORD', '')
PORT = os.environ.get('SOURCE_DB_PORT', '5432')

if not PASSWORD:
    print("警告：SOURCE_DB_PASSWORD 环境变量未设置！")
    exit(1)

CONN = psycopg2.connect(f'postgresql://{USER}:{PASSWORD}@{HOST}:{PORT}/{DATABASE}')

# 查询数据
df = pd.read_sql('''
        -- 步骤1：构建完整的时间维度（所有出现过的小时）
        with all_hours as (
            select date_trunc('hour', time_run - interval '15 minutes') as date_hour
            from region_spot_price_cleared  
            where time_run >= '2025-02-14'
            union
            select date_trunc('hour', time_run - interval '15 minutes') as date_hour
            from region_new_energy_prediction 
            where time_run >= '2025-02-14' and source_id = 'snv'
            union
            select time_run as date_hour
            from region_weather_prediction
            where time_run >= '2025-02-14'

            union
            select date_trunc('hour', time_run - interval '15 minutes') as date_hour
            from region_load_prediction
            where time_run >= '2025-02-14'
            union
            select date_trunc('hour', time_run - interval '15 minutes') as date_hour
            from region_out_prediction
            where time_run >= '2025-02-14'
            union
            select date_trunc('hour', time_run - interval '15 minutes') as date_hour
            from region_nonmarket_prediction 
            where time_run >= '2025-02-14'
            union
            select date_trunc('hour', time_run - interval '15 minutes') as date_hour
            from region_hydro_prediction
            where time_run >= '2025-02-14'
            union
            select date_trunc('hour', time_run - interval '15 minutes') as date_hour
            from region_power_prediction 
            where time_run >= '2025-02-14'
            union
            select date_trunc('hour', time_run - interval '15 minutes') as date_hour
            from region_balance_prediction
            where time_run >= '2025-02-14'

        ),

        -- 现货出清价格（原始粒度：15分钟；按小时聚合为平均值）
        spot_cleared as (
            select
                date_trunc('hour', time_run - interval '15 minutes') as date_hour,
                avg(price_real_time)  as price_real_time_avg,
                avg(price_day_ahead)  as price_day_ahead_avg
            from region_spot_price_cleared  
            where time_run >= '2025-02-14'
            group by 1
            order by date_hour desc
        ),

        -- 风光预测功率（原始粒度：15分钟；按小时聚合为平均值）
        wind_solar_prediction as (
            select
                date_trunc('hour', time_run - interval '15 minutes') as date_hour,
                avg(total_wind_solar) as total_wind_solar_pre_avg
            from region_new_energy_prediction 
            where time_run >= '2025-02-14'
              and source_id = 'snv'
            group by 1
            order by date_hour desc
        ),

        -- 天气预报（原始粒度：60分钟；按小时聚合）
        weather_prediction as (
            select
                time_run as date_hour,
                avg(case when elements = 'cloud' then value else null end) as cloud_avg,
                avg(case when elements = 'dir' then value else null end) as dir_avg,
                avg(case when elements = 'heat' then value else null end) as heat_avg,
                avg(case when elements = 'hum' then value else null end) as hum_avg,
                avg(case when elements = 'pm10' then value else null end) as pm10_avg,
                avg(case when elements = 'pm2_5' then value else null end) as pm2_5_avg,
                avg(case when elements = 'pre' then value else null end) as pre_avg,
                avg(case when elements = 'solar' then value else null end) as solar_radiation_avg,
                avg(case when elements = 'tem' then value else null end) as tem_avg,
                avg(case when elements = 'wind' then value else null end) as wind_speed_avg,
                avg(case when elements = 'wind10' then value else null end) as wind10_avg,
                avg(case when elements = 'wind100' then value else null end) as wind100_avg,
                avg(case when elements = 'wind120' then value else null end) as wind120_avg,
                avg(case when elements = 'wind180' then value else null end) as wind180_avg
            from region_weather_prediction
            where time_run >= '2025-02-14'
            group by 1
            order by date_hour desc
        )

        -- 区域负荷预测（原始粒度：15分钟；按小时聚合为平均值）
        ,load_prediction as (
            select
                date_trunc('hour', time_run - interval '15 minutes') as date_hour,
                avg(load) as load_pre_avg           -- 负荷功率预测值（MW），小时均值
            from region_load_prediction 
            where time_run >= '2025-02-14'
            and source_id = 'grid'
            group by date_hour
            order by date_hour desc
        )

        -- 外送功率预测（原始粒度：15分钟；按小时聚合为平均值）
        ,out_prediction as (
            select
                date_trunc('hour', time_run - interval '15 minutes') as date_hour,
                avg(out) as out_pre_avg          -- 外送功率预测值（MW），小时均值
            from region_out_prediction 
            where time_run >= '2025-02-14'
            and source_id = 'grid'
            group by date_hour
            order by date_hour desc
        )

        -- 非市场化机组出力预测（原始粒度：15分钟；按小时聚合为平均值）
        ,non_market_prediction as (
            select
                date_trunc('hour', time_run - interval '15 minutes') as date_hour,
                avg(nonmarket) non_market_pre_avg -- 未参与电力市场发电机组出力预测
            from region_nonmarket_prediction 
            where time_run >= '2025-02-14'
            and source_id = 'grid'
            group by date_hour
            order by date_hour desc
        )

        -- 水电出力预测（原始粒度：15分钟；按小时聚合为平均值）
        ,hydro_prediction as (
            select
                date_trunc('hour', time_run - interval '15 minutes') as date_hour,
                avg(hydro) hydro_pre_avg -- 水电出力预测
            from region_hydro_prediction 
            where time_run >= '2025-02-14'
            and source_id = 'grid'
            group by date_hour
            order by date_hour desc
        )

        -- 发电总出力预测（原始粒度：15分钟；按小时聚合为平均值）
        ,power_prediction as (
            select
                date_trunc('hour', time_run - interval '15 minutes') as date_hour,
                avg(power) power_pre_avg -- 总出力预测
            from region_power_prediction 
            where time_run >= '2025-02-14'
            and source_id = 'grid'
            group by date_hour
            order by date_hour desc
        )

        -- 供需差额预测（原始粒度：15分钟；按小时聚合为平均值）
        ,balance_prediction as (
            select
                date_trunc('hour', time_run - interval '15 minutes') as date_hour,
                avg(power) balance_pre_avg -- 供需差额预测
            from region_balance_prediction 
            where time_run >= '2025-02-14'
            and source_id = 'grid'
            group by date_hour
            order by date_hour desc
        )

        -- 主查询：以 all_hours 为主表，左连其他三张表
        select 
            h.date_hour,
            s.price_day_ahead_avg,
            w.total_wind_solar_pre_avg,
            we.cloud_avg,
            we.dir_avg,
            we.heat_avg,
            we.hum_avg,
            we.pm10_avg,
            we.pm2_5_avg,
            we.pre_avg,
            we.solar_radiation_avg,
            we.tem_avg,
            we.wind_speed_avg,
            we.wind10_avg,
            we.wind100_avg,
            we.wind120_avg,
            we.wind180_avg,
            lp.load_pre_avg,
            op.out_pre_avg,
            nmp.non_market_pre_avg,
            hp.hydro_pre_avg,
            pop.power_pre_avg,
            bp.balance_pre_avg
        from all_hours h
        left join spot_cleared s on h.date_hour = s.date_hour
        left join wind_solar_prediction w on h.date_hour = w.date_hour
        left join weather_prediction we on h.date_hour = we.date_hour

        left join load_prediction lp on h.date_hour = lp.date_hour
        left join out_prediction op on h.date_hour = op.date_hour
        left join non_market_prediction nmp on h.date_hour = nmp.date_hour
        left join hydro_prediction hp on h.date_hour = hp.date_hour
        left join power_prediction pop on h.date_hour = pop.date_hour
        left join balance_prediction bp on h.date_hour = bp.date_hour
        order by h.date_hour desc;
    ''', CONN)
print(df.head())

CONN.close()

# --- 第二部分：定义目标数据库连接 ---
# 从环境变量读取目标数据库配置
target_username = os.environ.get('DB_USER', 'seniverse')
target_password = os.environ.get('DB_PASSWORD', '')
target_host = os.environ.get('DB_HOST', 'pgm-bp1q60i1ca0xnwv98o.pg.rds.aliyuncs.com')
target_port = os.environ.get('DB_PORT', '1921')
target_database = os.environ.get('DB_NAME', 'bopha')
table_name = 'anhui_hourly_spot_cleared_wide'

if not target_password:
    print("警告：DB_PASSWORD 环境变量未设置！")
    exit(1)

# 创建 SQLAlchemy 引擎用于 to_sql
engine_url = f'postgresql+psycopg2://{target_username}:{target_password}@{target_host}:{target_port}/{target_database}'
engine = create_engine(engine_url)

# 创建 psycopg2 连接用于执行 DDL
conn_str = f"dbname={target_database} user={target_username} password={target_password} host={target_host} port={target_port}"
conn = psycopg2.connect(conn_str)
cursor = conn.cursor()

try:
    cursor.execute(f"DROP TABLE IF EXISTS {table_name};")
    
    create_table_sql = f"""
    CREATE TABLE {table_name} (
        date_hour               TIMESTAMPTZ PRIMARY KEY,
        price_day_ahead_avg     NUMERIC(10, 4),
        total_wind_solar_pre_avg NUMERIC(10, 2),
        cloud_avg               NUMERIC(10, 2),
        dir_avg                 NUMERIC(10, 2),
        heat_avg                NUMERIC(10, 2),
        hum_avg                 NUMERIC(10, 2),
        pm10_avg                NUMERIC(10, 2),
        pm2_5_avg               NUMERIC(10, 2),
        pre_avg                 NUMERIC(10, 2),
        solar_radiation_avg     NUMERIC(10, 2),
        tem_avg                 NUMERIC(10, 2),
        wind_speed_avg          NUMERIC(10, 2),
        wind10_avg              NUMERIC(10, 2),
        wind100_avg             NUMERIC(10, 2),
        wind120_avg             NUMERIC(10, 2),
        wind180_avg             NUMERIC(10, 2),
        load_pre_avg            NUMERIC(10, 2),
        out_pre_avg             NUMERIC(10, 2),
        non_market_pre_avg      NUMERIC(10, 2),
        hydro_pre_avg           NUMERIC(10, 2),
        power_pre_avg           NUMERIC(10, 2),
        balance_pre_avg         NUMERIC(10, 2)
    );
    """
    cursor.execute(create_table_sql)
    print(f"表 '{table_name}' 创建成功。")

    comments_sql = [
        f"COMMENT ON COLUMN {table_name}.date_hour IS '统计小时点 (截断至整点)';",
        f"COMMENT ON COLUMN {table_name}.price_day_ahead_avg IS '日前出清均价 (元/MWh)';",
        f"COMMENT ON COLUMN {table_name}.total_wind_solar_pre_avg IS '风光预测总功率均值 (MW)';",
        f"COMMENT ON COLUMN {table_name}.tem_avg IS '平均气温 (℃)';",
        f"COMMENT ON COLUMN {table_name}.load_pre_avg IS '区域负荷预测均值 (MW)';",
        f"COMMENT ON COLUMN {table_name}.balance_pre_avg IS '供需差额预测均值 (MW)';",
    ]
    
    for sql in comments_sql:
        cursor.execute(sql)
    
    conn.commit()
    print("列备注添加成功。")

    df.to_sql(
        name=table_name,
        con=engine,
        if_exists='append',
        index=False,
        method='multi',
        chunksize=1000
    )
    
    print(f"数据成功追加到表 '{table_name}'，共 {len(df)} 条记录。")

except Exception as e:
    conn.rollback()
    print(f"发生错误: {e}")
finally:
    cursor.close()
    conn.close()
    engine.dispose()
